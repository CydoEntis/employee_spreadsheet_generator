import os
import shutil
import openpyxl
import csv
import argparse
from datetime import datetime, timedelta

TEMPLATE_DIR = "templates"
TEMPLATE_FILE = "template.xlsx"
OUTPUT_DIR = "spreadsheets"
PAYROLL_DATE_FORMAT = "%m/%d/%Y"


# Get the root directory of the script
root_dir = os.path.dirname(os.path.abspath(__file__))

# Define paths relative to the script's location
template_path = os.path.join(root_dir, TEMPLATE_DIR, TEMPLATE_FILE)
output_path = os.path.join(root_dir, OUTPUT_DIR)


def read_employee_data(file_path):
    """Reads employee data from a CSV file and returns a list of dictionaries."""
    employees = []
    with open(file_path, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            employees.append({"id": row["ID"], "name": row["Name"]})
    return employees


def get_friday_of_current_week():
    """Finds the Friday of the current week."""
    today = datetime.today()
    days_until_friday = (4 - today.weekday()) % 7  # Friday is weekday 4
    return today + timedelta(days=days_until_friday)


def calculate_pay_and_deal_periods():
    """Calculates pay period (last week Mon-Sat) and deal period (two weeks ago Mon-Sat)."""
    period_end = get_friday_of_current_week()

    # Pay period: Last week's Monday - Saturday
    pay_period_start = period_end - timedelta(days=11)
    pay_period_end = period_end - timedelta(days=6)

    # Deal period: Two weeks back, Monday - Saturday
    deal_period_start = period_end - timedelta(days=18)
    deal_period_end = period_end - timedelta(days=13)

    return {
        "pay_period": f"{pay_period_start.strftime('%m/%d/%Y')} to {pay_period_end.strftime('%m/%d/%Y')}",
        "deal_period": f"{deal_period_start.strftime('%m/%d/%Y')} to {deal_period_end.strftime('%m/%d/%Y')}",
        "paid_date": period_end.strftime('%m/%d/%Y'),
        "file_date": pay_period_start.strftime('%Y-%m-%d'),  # Now uses pay period start date
    }


def clear_directory(directory):
    """Clears all files and subdirectories in the given directory."""
    if os.path.exists(directory):
        shutil.rmtree(directory)
    os.makedirs(directory, exist_ok=True)


def generate_excel_for_employee(employee, periods):
    """Generates an Excel payroll file for a given employee."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active  # Assuming the relevant sheet is the first one

    # Update relevant fields
    ws["B1"] = employee["id"]  # Employee ID
    ws["B2"] = employee["name"]  # Employee Name
    ws["B3"] = periods["paid_date"]  # Paid Date
    ws["B4"] = periods["pay_period"]  # Pay Period
    ws["B5"] = periods["deal_period"]  # Deal Period

    # Generate the output file name
    first_name, last_name = employee["name"].split(' ', 1)
    output_filename = f"{first_name}_{last_name}_{periods['file_date']}.xlsx"

    # Save the new Excel file
    file_path = os.path.join(OUTPUT_DIR, output_filename)
    wb.save(file_path)


def main():
    """Main function to generate employee spreadsheets."""
    # Set up argument parsing
    parser = argparse.ArgumentParser(description="Generate spreadsheets for employees.")
    parser.add_argument("employee_data_file", help="Path to the CSV file containing employee data.")
    args = parser.parse_args()

    # Read employee data
    employees = read_employee_data(args.employee_data_file)

    # Clear the spreadsheets directory
    clear_directory(output_path)

    # Calculate periods
    periods = calculate_pay_and_deal_periods()

    # Process each employee
    for emp in employees:
        generate_excel_for_employee(emp, periods)

    print("Excel files generated successfully!")


if __name__ == "__main__":
    main()
